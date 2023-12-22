import pandas as pd
import shutil
import os
from openpyxl import load_workbook
import tkinter as tk

class Modelo:
    def __init__(self, vista):
        self.archivo_subido = None
        self.archivo_standard = "recursos/planilla_standard.xlsx"
        self.columna_standard_index = 2
        self.columnas_a_insertar = [8, 9, 10, 11, 12, 13, 14, 15, 17, 18, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 39, 40, 41, 42, 43, 44, 45, 47, 50, 51, 52, 53, 54, 55, 56, 60, 61, 62, 63, 64, 65]
        self.columnas_a_insertar_tipo2 = [4, 5, 6, 7, 8, 9, 10, 11, 13, 14, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 35, 36, 37, 38, 39, 40, 41, 43, 46, 47, 48, 49, 50, 51, 52, 56, 57, 58, 59, 60, 61, 63 ]
        self.filas_a_pegar = [6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 37, 38, 39, 40, 41, 42, 43, 45, 48, 49, 50, 51, 52, 53, 54, 58, 59, 60, 61, 62, 63, 65]
        self.columnas_a_insertarR = [7, 9, 13, 17, 21, 23, 25, 27, 29, 31, 33, 39]
        self.filas_a_pegarR = [5, 7, 11, 12, 15, 17, 19, 21, 23, 25, 27, 31]
        self.columnas_a_insertarR2 = [3, 5, 9, 10, 13, 15, 17, 19, 21, 23, 25, 29]
        self.vista = vista
        self.nombres_filas = {1: "Efectivo y Equivalentes al Efectivo", 2: "Otra Fila", 3: "Otra Fila Más"}
        self.nombres_columnas = {1: "Chilefilms", 20: "Otra Columna", 30: "Otra Columna Más"}
        
    def cargar_archivo(self, ruta_archivo):
        try:
            
            self.archivo_subido = pd.ExcelFile(ruta_archivo)
        except Exception as e:
            print(f"Error al cargar el archivo: {e}")

    def buscar_total_tipo1(self):
        if self.archivo_subido is not None:
            if "Estado" in self.archivo_subido.sheet_names:
                columna_subida = self.archivo_subido.parse("Estado").iloc[:, 2]  # Columna ubicada en el índice 3 (columna C en Excel)
                return columna_subida[self.columnas_a_insertar]
            else:
                print("No se encontró la hoja 'Estado' en el archivo subido.")
        else:
            print("No se ha cargado ningún archivo.")

    def buscar_total_tipo2(self):
        if self.archivo_subido is not None:
            if "Estado" in self.archivo_subido.sheet_names:
                columna_subida = self.archivo_subido.parse("Estado").iloc[:, 2]  # Columna ubicada en el índice 3 (columna C en Excel)
                return columna_subida[self.columnas_a_insertar_tipo2]
            else:
                print("No se encontró la hoja 'Estado' en el archivo subido.")
        else:
            print("No se ha cargado ningún archivo.")

    def buscar_total_tipo3(self):
        if self.archivo_subido is not None:
            if "Estado" in self.archivo_subido.sheet_names:
                columna_subida = self.archivo_subido.parse("Estado").iloc[:, 24]  # Columna ubicada en el índice 25 (columna Y en Excel)
                return columna_subida[self.columnas_a_insertar_tipo2]
            else:
                print("No se encontró la hoja 'Estado' en el archivo subido.")
        else:
            print("No se ha cargado ningún archivo.")

    def buscar_total_tipo4(self):
        if self.archivo_subido is not None:
            if "Estado$" in self.archivo_subido.sheet_names:
                columna_subida = self.archivo_subido.parse("Estado$").iloc[:, 28]  # Columna ubicada en el índice 25 (columna Y en Excel)
                return columna_subida[self.columnas_a_insertar_tipo2]
            else:
                print("No se encontró la hoja 'Estado$' en el archivo subido.")
        else:
            print("No se ha cargado ningún archivo.")
   
    def insertar_en_standard_tipo1(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 1).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegar, self.columnas_a_insertar, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 1}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 1, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo standard para Tipo 1.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")
                
           
            self.vista.mostrar_proceso_finalizado(1)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para Tipo 1: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_tipo2(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 2).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegar, self.columnas_a_insertar_tipo2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 2}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 2, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo estándar para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_tipo3(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 3).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegar, self.columnas_a_insertar_tipo2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 3}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 3, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo standard para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_tipo4(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 4).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegar, self.columnas_a_insertar_tipo2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 4}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 4, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo estándar para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado(4)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)
    
    def insertar_en_standard_tipo5(self, total_chilefilms):
        
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 5).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegar, self.columnas_a_insertar_tipo2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 5}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 5, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo standard para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_tipo6(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 6).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegar, self.columnas_a_insertar_tipo2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 6}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 6, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo estándar para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_tipo7(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 7).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegar, self.columnas_a_insertar_tipo2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 7}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 7, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo estándar para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_tipo8(self, total_chilefilms):
        
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 8).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegar, self.columnas_a_insertar_tipo2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 8}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 8, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo standard para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)
    
    def insertar_valor_en_celda(self, fila, columna, valor):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]

                estilo_celda = hoja_estado.cell(row=fila, column=columna)._style
                hoja_estado.cell(row=fila, column=columna, value=valor)._style = estilo_celda

                archivo_standard.save(self.archivo_standard)

                
                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)
                shutil.copy(self.archivo_standard, ruta_copia)
                print(f"Copia guardada en el escritorio como {nombre_copia}")

                nombre_fila = self.nombres_filas.get(fila, f"Fila {fila}")
                nombre_columna = self.nombres_columnas.get(columna, f"Columna {columna}")
                print(f"Valor '{valor}' insertado en {nombre_fila}, {nombre_columna} correctamente.")
                return True
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")
                return False

        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar: {e}"
            print(mensaje_error)
            return False
    
    def buscar_total_resultado(self):
        if self.archivo_subido is not None:
            if "Resultado" in self.archivo_subido.sheet_names:
                columna_subida = self.archivo_subido.parse("Resultado").iloc[:, 2]  # Columna ubicada en el índice 3 (columna C en Excel)
                return columna_subida[self.columnas_a_insertarR]
            else:
                print("No se encontró la hoja 'Estado' en el archivo subido.")
        else:
            print("No se ha cargado ningún archivo.")

    def buscar_total_resultado2(self):
        if self.archivo_subido is not None:
            if "Resultado" in self.archivo_subido.sheet_names:
                columna_subida = self.archivo_subido.parse("Resultado").iloc[:, 2]  # Columna ubicada en el índice 3 (columna C en Excel)
                return columna_subida[self.columnas_a_insertarR2]
            else:
                print("No se encontró la hoja 'Estado' en el archivo subido.")
        else:
            print("No se ha cargado ningún archivo.")

    def buscar_total_resultado3(self):
        if self.archivo_subido is not None:
            if "Resultado" in self.archivo_subido.sheet_names:
                columna_subida = self.archivo_subido.parse("Resultado").iloc[:, 24]  # Columna ubicada en el índice 25 (columna Y en Excel)
                return columna_subida[self.columnas_a_insertarR2]
            else:
                print("No se encontró la hoja 'Estado' en el archivo subido.")
        else:
            print("No se ha cargado ningún archivo.")

    def buscar_total_resultado4(self):
        if self.archivo_subido is not None:
            if "Resultado$" in self.archivo_subido.sheet_names:
                columna_subida = self.archivo_subido.parse("Resultado$").iloc[:, 28]  # Columna ubicada en el índice 25 (columna Y en Excel)
                return columna_subida[self.columnas_a_insertarR2]
            else:
                print("No se encontró la hoja 'Estado$' en el archivo subido.")
        else:
            print("No se ha cargado ningún archivo.")       

    def insertar_en_standard_Resul(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Resultado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 1).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegarR, self.columnas_a_insertarR, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 1}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 1, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo standard.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")
                
           
            self.vista.mostrar_proceso_finalizado2(1)
        except Exception as e:
            mensaje_error = f"Error al insertar: {e}"
            self.vista.mostrar_error(mensaje_error)
   
    def insertar_en_standard_Resul2(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Resultado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 2).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegarR, self.columnas_a_insertarR2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 2}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 2, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo estándar para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado2(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_Resul3(self, total_chilefilms):
    
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Resultado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 3).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegarR, self.columnas_a_insertarR2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 3}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 3, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo standard para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado2(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_Resul4(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Resultado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 4).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegarR, self.columnas_a_insertarR2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 4}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 4, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo estándar para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado2(4)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_Resul5(self, total_chilefilms):
        
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Resultado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 5).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegarR, self.columnas_a_insertarR2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 5}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 5, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo standard para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado2(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_Resul6(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Resultado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 6).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegarR, self.columnas_a_insertarR2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 6}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 6, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo estándar para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado2(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_Resul7(self, total_chilefilms):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Resultado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 7).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegarR, self.columnas_a_insertarR2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 7}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 7, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo estándar para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado2(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def insertar_en_standard_Resul8(self, total_chilefilms):
        
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Resultado"]

                if not hoja_estado.sheet_state == "visible":
                    hoja_estado.sheet_state = "visible"

                columna_standard = [hoja_estado.cell(row=fila, column=self.columna_standard_index + 8).value for fila in
                                    range(1, hoja_estado.max_row + 1)]

                for fila_paste, fila_copy, valor in zip(self.filas_a_pegarR, self.columnas_a_insertarR2, total_chilefilms):
                    print(f"Copiando de fila {fila_copy}, pegando en fila {fila_paste}, columna {self.columna_standard_index + 8}, valor: {valor}")
                    hoja_estado.cell(row=fila_paste, column=self.columna_standard_index + 8, value=valor)

                archivo_standard.save(self.archivo_standard)

                print(f"Datos insertados correctamente en el archivo standard para Chilefilms.")

                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)

                shutil.copy(self.archivo_standard, ruta_copia)

                print(f"Copia guardada en el escritorio como {nombre_copia}")
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")

            self.vista.mostrar_proceso_finalizado2(2)
        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar para CHF Internacional: {e}"
            self.vista.mostrar_error(mensaje_error)

    def limpiar_planilla(self):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Estado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Estado"]
                filas_a_limpiar_estado = [6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 37, 38, 39, 40, 41, 42, 43, 45, 48, 49, 50, 51, 52, 53, 54]
                self.limpiar_celdas(hoja_estado, filas_a_limpiar_estado)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_resultado = archivo_standard["Resultado"]
                filas_a_limpiar_resultado = [5, 7, 11, 12, 15, 17, 19, 21, 23, 25, 27, 31]
                self.limpiar_celdas(hoja_resultado, filas_a_limpiar_resultado)

            archivo_standard.save(self.archivo_standard)
            return True
        except Exception as e:
            mensaje_error = f"Error al limpiar la planilla: {e}"
            print(mensaje_error)
            raise RuntimeError(mensaje_error)

    def limpiar_celdas(self, hoja, filas_a_limpiar=None):
        columnas_a_limpiar = [3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17, 18, 19]
        
        if filas_a_limpiar is None:
            filas_a_limpiar = []

        for columna in columnas_a_limpiar:
            for fila in filas_a_limpiar:
                hoja.cell(row=fila, column=columna).value = None

    def descargar_manual(self):
        try:
            shutil.copy("recursos/manual.pdf", os.path.join(os.path.expanduser("~"), "Desktop", "manual.pdf"))
        except Exception as e:
            mensaje_error = f"Error al descargar el manual: {e}"
            print(mensaje_error)
            raise RuntimeError(mensaje_error)
        
    def insertar_valor_en_celda_R(self, fila, columna, valor):
        try:
            archivo_standard = load_workbook(self.archivo_standard)

            if "Resultado" in archivo_standard.sheetnames:
                hoja_estado = archivo_standard["Resultado"]

                estilo_celda = hoja_estado.cell(row=fila, column=columna)._style
                hoja_estado.cell(row=fila, column=columna, value=valor)._style = estilo_celda

                archivo_standard.save(self.archivo_standard)

                
                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
                nombre_copia = "planilla_standard_modificada.xlsx"
                ruta_copia = os.path.join(escritorio, nombre_copia)
                shutil.copy(self.archivo_standard, ruta_copia)
                print(f"Copia guardada en el escritorio como {nombre_copia}")

                nombre_fila = self.nombres_filas.get(fila, f"Fila {fila}")
                nombre_columna = self.nombres_columnas.get(columna, f"Columna {columna}")
                print(f"Valor '{valor}' insertado en {nombre_fila}, {nombre_columna} correctamente.")
                return True
            else:
                print("No se encontró la hoja 'Estado' en el archivo estándar.")
                return False

        except Exception as e:
            mensaje_error = f"Error al insertar en el archivo estándar: {e}"
            print(mensaje_error)
            return False